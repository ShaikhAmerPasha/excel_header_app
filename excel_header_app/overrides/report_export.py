import datetime
import io
import json

import frappe
from frappe import _
from frappe.utils import cint


def _get_default_company() -> str:
    try:
        company = (
            frappe.defaults.get_defaults().get("company")
            or frappe.db.get_single_value("Global Defaults", "default_company")
        )
        return company or "Company"
    except Exception:
        return "Company"


def _today_formatted() -> str:
    return datetime.date.today().strftime("%d-%m-%Y")


def _build_xlsx_with_header(data: list, report_name: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (report_name or "Report")[:31]

    num_cols = len(data[0]) if data else 1
    last_col = get_column_letter(num_cols)

    company_font = Font(name="Calibri", bold=True,   size=14)
    title_font   = Font(name="Calibri", bold=True,   size=12)
    meta_font    = Font(name="Calibri", italic=True, size=10)
    header_font  = Font(name="Calibri", bold=True,   size=10, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="4472C4")

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    thin        = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([_get_default_company()])
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].font      = company_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    ws.append([report_name or "Report"])
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].font      = title_font
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    ws.append([f"Downloaded On: {_today_formatted()}"])
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].font      = meta_font
    ws["A3"].alignment = center
    ws.row_dimensions[3].height = 16

    ws.append([""])
    ws.row_dimensions[4].height = 8

    if data:
        ws.append(data[0])
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = cell_border
        ws.row_dimensions[5].height = 15

        for row_data in data[1:]:
            safe_row = []
            for val in row_data:
                if isinstance(val, datetime.timedelta):
                    safe_row.append(str(val))
                else:
                    safe_row.append(val)
            ws.append(safe_row)

        for row in ws.iter_rows(
            min_row=6, max_row=ws.max_row, min_col=1, max_col=num_cols
        ):
            for cell in row:
                cell.border    = cell_border
                cell.alignment = left

    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(5, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len * 1.2, 10), 60)

    ws.freeze_panes = "A6"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@frappe.whitelist()
def export_query():
    form_params      = frappe.form_dict
    file_format_type = form_params.get("file_format_type", "Excel")
    report_name      = form_params.get("report_name", "")
    filters          = form_params.get("filters")
    visible_idx      = form_params.get("visible_idx")

    if file_format_type != "Excel":
        from frappe.desk.query_report import export_query as _core
        return _core()

    report_doc = frappe.get_doc("Report", report_name)
    if report_doc.report_type == "Report Builder":
        from frappe.desk.reportview import export_query as rb_export
        return rb_export()

    if filters and isinstance(filters, str):
        filters = json.loads(filters)
    if visible_idx and isinstance(visible_idx, str):
        visible_idx = json.loads(visible_idx)

    from frappe.desk.query_report import run as run_report

    run_result = run_report(
        report_name,
        filters=filters,
        ignore_prepared_report=True,
        are_default_filters=False,
    )

    columns = run_result.get("columns") or []
    result  = run_result.get("result")  or []

    col_labels = []
    for col in columns:
        if isinstance(col, dict):
            col_labels.append(col.get("label") or col.get("fieldname") or "")
        else:
            col_labels.append(str(col))

    if visible_idx:
        visible_idx = [cint(i) for i in visible_idx]
        col_labels = [col_labels[i] for i in visible_idx if i < len(col_labels)]

    data_rows = []
    for row in result:
        if isinstance(row, dict):
            values = list(row.values())
        elif isinstance(row, (list, tuple)):
            values = list(row)
        else:
            values = [row]

        if visible_idx:
            values = [values[i] for i in visible_idx if i < len(values)]

        data_rows.append(values)

    all_data   = [col_labels] + data_rows
    xlsx_bytes = _build_xlsx_with_header(all_data, report_name)

    frappe.response["filename"]    = f"{report_name}.xlsx"
    frappe.response["filecontent"] = xlsx_bytes
    frappe.response["type"]        = "binary"
    frappe.response["doctype"]     = None